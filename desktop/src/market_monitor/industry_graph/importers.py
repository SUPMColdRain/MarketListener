"""Evidence importers for HTML, Excel, PDF and announcement sources.

FULL-701 keeps every imported record traceable to the original file: each
record carries the source id, source type, SHA-256 of the original file, the
parser version, the extraction time and a precise location (DOM path, cell,
page+line or line+offset).  Re-importing the same file with the same parser
version is idempotent and returns no duplicate records.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from .models import EvidenceLocation

PARSED_VERSION = "1.0.0"


class GraphImportError(ValueError):
    """A classified import failure (format, content or parsing)."""

    def __init__(self, category: str, message: str) -> None:
        super().__init__(message)
        self.category = category
        self.message = message


@dataclass(frozen=True)
class ImportedRecord:
    """One extracted snippet plus its precise original-file location."""

    record_id: str
    source_id: str
    source_type: str
    location: EvidenceLocation
    snippet: str
    sha256: str
    parsed_version: str
    extracted_at: str

    def to_evidence(self, evidence_id: str) -> dict[str, object]:
        document: dict[str, object] = {
            "schema_version": 1,
            "evidence_id": evidence_id,
            "source_id": self.source_id,
            "source_type": self.source_type,
            "location": self.location.as_mapping(),
            "parsed_version": self.parsed_version,
            "extracted_at": self.extracted_at,
            "sha256": self.sha256,
        }
        return document


@dataclass(frozen=True)
class ImportResult:
    """Outcome of one import call, including idempotent duplicates."""

    source_id: str
    source_type: str
    sha256: str
    parsed_version: str
    extracted_at: str
    records: tuple[ImportedRecord, ...]
    duplicate: bool = False


class GraphImporter:
    """Format-dispatching importer with per-run duplicate detection."""

    def __init__(self, parsed_version: str = PARSED_VERSION, now: datetime | None = None) -> None:
        self.parsed_version = parsed_version
        self._now = now or datetime.now(timezone.utc)
        self._seen: set[tuple[str, str]] = set()

    def import_file(self, path: Path, source_type: str | None = None) -> ImportResult:
        resolved = Path(path)
        if not resolved.is_file():
            raise GraphImportError("NOT_FOUND", f"source file not found: {resolved}")
        raw = resolved.read_bytes()
        sha256 = hashlib.sha256(raw).hexdigest()
        chosen_type = source_type or _source_type_for_suffix(resolved.suffix)
        key = (sha256, self.parsed_version)
        if key in self._seen:
            return ImportResult(
                source_id=_source_id(resolved),
                source_type=chosen_type,
                sha256=sha256,
                parsed_version=self.parsed_version,
                extracted_at=_iso(self._now),
                records=(),
                duplicate=True,
            )
        self._seen.add(key)
        records = _extract(chosen_type, resolved, raw, sha256, self.parsed_version, self._now)
        return ImportResult(
            source_id=_source_id(resolved),
            source_type=chosen_type,
            sha256=sha256,
            parsed_version=self.parsed_version,
            extracted_at=_iso(self._now),
            records=tuple(records),
        )


def _source_type_for_suffix(suffix: str) -> str:
    normalized = suffix.lower()
    if normalized in {".html", ".htm"}:
        return "HTML"
    if normalized in {".xlsx", ".xls"}:
        return "EXCEL"
    if normalized == ".pdf":
        return "PDF"
    if normalized in {".txt", ".md", ".text"}:
        return "ANNOUNCEMENT"
    raise GraphImportError("UNSUPPORTED_FORMAT", f"unsupported source suffix: {suffix}")


def _source_id(path: Path) -> str:
    resolved = str(path.resolve())
    digest = hashlib.sha1(resolved.encode("utf-8")).hexdigest()[:8]
    value = path.stem.strip().lower()
    value = re.sub(r"[^a-z0-9_.-]+", "_", value).strip("._")
    return f"{value or 'source'}.{digest}"


def _iso(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.isoformat()


def _extract(
    source_type: str,
    path: Path,
    raw: bytes,
    sha256: str,
    parsed_version: str,
    now: datetime,
) -> list[ImportedRecord]:
    if source_type == "HTML":
        return _extract_html(path, raw, sha256, parsed_version, now)
    if source_type == "EXCEL":
        return _extract_excel(path, raw, sha256, parsed_version, now)
    if source_type == "PDF":
        return _extract_pdf(path, raw, sha256, parsed_version, now)
    if source_type in {"ANNOUNCEMENT", "TEXT"}:
        return _extract_text(path, raw, sha256, parsed_version, now, source_type)
    raise GraphImportError("UNSUPPORTED_FORMAT", f"unsupported source type: {source_type}")


def _extract_html(
    path: Path,
    raw: bytes,
    sha256: str,
    parsed_version: str,
    now: datetime,
) -> list[ImportedRecord]:
    try:
        from bs4 import BeautifulSoup
    except ImportError as error:  # pragma: no cover - environment guard
        raise GraphImportError("PARSER_MISSING", "beautifulsoup4 is required for HTML import") from error
    try:
        text = raw.decode("utf-8")
        soup = BeautifulSoup(text, "html.parser")
        records: list[ImportedRecord] = []
        for element in soup.find_all(["td", "th", "p", "li"]):
            snippet = _clean(element.get_text(" ", strip=True))
            if not snippet:
                continue
            dom = _dom_path(element)
            records.append(
                ImportedRecord(
                    record_id=f"{_source_id(path)}.{len(records) + 1:04d}",
                    source_id=_source_id(path),
                    source_type="HTML",
                    location=EvidenceLocation(dom=dom),
                    snippet=snippet,
                    sha256=sha256,
                    parsed_version=parsed_version,
                    extracted_at=_iso(now),
                )
            )
    except (UnicodeDecodeError, ValueError) as error:
        raise GraphImportError("ENCODING", f"HTML decoding failed: {error}") from error
    _require_records(records, "HTML")
    return records


def _extract_excel(
    path: Path,
    raw: bytes,
    sha256: str,
    parsed_version: str,
    now: datetime,
) -> list[ImportedRecord]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - environment guard
        raise GraphImportError("PARSER_MISSING", "openpyxl is required for Excel import") from error
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        records: list[ImportedRecord] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    snippet = _clean(str(cell.value)) if cell.value is not None else ""
                    if not snippet:
                        continue
                    records.append(
                        ImportedRecord(
                            record_id=f"{_source_id(path)}.{len(records) + 1:04d}",
                            source_id=_source_id(path),
                            source_type="EXCEL",
                            location=EvidenceLocation(cell=f"{sheet.title}!{cell.coordinate}"),
                            snippet=snippet,
                            sha256=sha256,
                            parsed_version=parsed_version,
                            extracted_at=_iso(now),
                        )
                    )
        workbook.close()
    except Exception as error:
        raise GraphImportError("CORRUPT_FILE", f"Excel parsing failed: {error}") from error
    _require_records(records, "EXCEL")
    return records


def _extract_pdf(
    path: Path,
    raw: bytes,
    sha256: str,
    parsed_version: str,
    now: datetime,
) -> list[ImportedRecord]:
    try:
        from pypdf import PdfReader
    except ImportError as error:  # pragma: no cover - environment guard
        raise GraphImportError("PARSER_MISSING", "pypdf is required for PDF import") from error
    try:
        reader = PdfReader(path)
        records: list[ImportedRecord] = []
        for page_number, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            for line_number, line in enumerate(text.splitlines(), start=1):
                snippet = _clean(line)
                if not snippet:
                    continue
                records.append(
                    ImportedRecord(
                        record_id=f"{_source_id(path)}.{len(records) + 1:04d}",
                        source_id=_source_id(path),
                        source_type="PDF",
                        location=EvidenceLocation(page=page_number, line=line_number),
                        snippet=snippet,
                        sha256=sha256,
                        parsed_version=parsed_version,
                        extracted_at=_iso(now),
                    )
                )
    except Exception as error:
        raise GraphImportError("CORRUPT_FILE", f"PDF parsing failed: {error}") from error
    _require_records(records, "PDF")
    return records


def _extract_text(
    path: Path,
    raw: bytes,
    sha256: str,
    parsed_version: str,
    now: datetime,
    source_type: str,
) -> list[ImportedRecord]:
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as error:
        raise GraphImportError("ENCODING", f"text decoding failed: {error}") from error
    records: list[ImportedRecord] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        snippet = _clean(line)
        if not snippet:
            continue
        offset = text.find(line) if line else 0
        records.append(
            ImportedRecord(
                record_id=f"{_source_id(path)}.{len(records) + 1:04d}",
                source_id=_source_id(path),
                source_type=source_type,
                location=EvidenceLocation(line=line_number, offset=max(offset, 0)),
                snippet=snippet,
                sha256=sha256,
                parsed_version=parsed_version,
                extracted_at=_iso(now),
            )
        )
    _require_records(records, source_type)
    return records


def _dom_path(element) -> str:
    parts: list[str] = []
    current = element
    while current is not None and current.name is not None:
        name = current.name
        index = 1
        sibling = current.previous_sibling
        while sibling is not None:
            if getattr(sibling, "name", None) == name:
                index += 1
            sibling = sibling.previous_sibling
        parts.append(f"{name}[{index}]")
        current = current.parent
    return "/".join(reversed(parts))


def _clean(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _require_records(records: Iterable[ImportedRecord], source_type: str) -> None:
    if not any(records):
        raise GraphImportError("EMPTY", f"{source_type} source contains no extractable text")
